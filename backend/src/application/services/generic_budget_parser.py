import openpyxl

TOTAL_KEYWORDS = [
    'total', 'total general', 'grand total', 'suma total',
    'valor total', 'total proyecto', 'total oferta',
    'total costo', 'total presupuesto', 'subtotal'
]

def detectar_total_venta(wb: openpyxl.Workbook) -> float:
    """
    Estrategia 1: para LYRA/CARSAN (si tiene Hoja2), buscar el valor máximo en la Columna C (col 3) de Hoja2.
    Estrategia 2: buscar fila con keyword de total → valor numérico más grande de esa fila (ignorando secciones parciales).
    Estrategia 3: si no hay keyword → valor numérico máximo de toda la hoja principal.
    """
    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    if 'hoja2' in sheet_names_lower:
        hoja2 = wb[wb.sheetnames[sheet_names_lower.index('hoja2')]]
        c_vals = []
        for r in range(1, hoja2.max_row + 1):
            val = hoja2.cell(r, 3).value
            if isinstance(val, (int, float)) and val > 0:
                c_vals.append(float(val))
        if c_vals:
            return max(c_vals)

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    if "subtotal (a)" in cell.lower() or "subtotal (b)" in cell.lower():
                        continue
                    if any(kw in cell.lower() for kw in TOTAL_KEYWORDS):
                        nums = [v for v in row if isinstance(v, (int, float)) and v > 0]
                        if nums:
                            return float(max(nums))

    # Estrategia 3: máximo numérico de la hoja principal
    sheet = wb.active
    max_val = 0.0
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, (int, float)) and cell > max_val:
                max_val = float(cell)

    if max_val == 0:
        raise ValueError(
            "No se detectó un total en el Excel de Venta. "
            "Asegúrate de que el archivo contenga una fila con 'Total' y un valor numérico."
        )
    return max_val


def detectar_total_costo(wb: openpyxl.Workbook) -> float:
    """
    Mismo formato que Venta pero con columna de costo separada.
    Estrategia 1: para LYRA/CARSAN (si tiene Hoja2), buscar el valor máximo en la Columna L (col 12) de Hoja2.
    Estrategia 2: buscar keyword de total → si hay múltiples columnas numéricas
    en esa fila, tomar la segunda (col1=Venta, col2=Costo).
    Si solo hay una columna numérica → usarla como costo.
    """
    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    if 'hoja2' in sheet_names_lower:
        hoja2 = wb[wb.sheetnames[sheet_names_lower.index('hoja2')]]
        if hoja2.max_column >= 12:
            l_vals = []
            for r in range(1, hoja2.max_row + 1):
                val = hoja2.cell(r, 12).value
                if isinstance(val, (int, float)) and val > 10.0:  # Excluir porcentajes pequeños
                    l_vals.append(float(val))
            if l_vals:
                return max(l_vals)
        c_vals = []
        for r in range(1, hoja2.max_row + 1):
            val = hoja2.cell(r, 3).value
            if isinstance(val, (int, float)) and val > 0:
                c_vals.append(float(val))
        if c_vals:
            return max(c_vals)

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    if "subtotal (a)" in cell.lower() or "subtotal (b)" in cell.lower():
                        continue
                    if any(kw in cell.lower() for kw in TOTAL_KEYWORDS):
                        nums = [v for v in row if isinstance(v, (int, float)) and v > 0]
                        if len(nums) >= 2:
                            return float(nums[1])  # segunda columna = Costo
                        elif len(nums) == 1:
                            return float(nums[0])  # única columna disponible

    raise ValueError(
        "No se detectó un total de costo en el Excel. "
        "Asegúrate de que el archivo contenga una fila 'Total' con columna de costo."
    )
