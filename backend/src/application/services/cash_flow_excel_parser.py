"""
Parser para el Excel "Flujo de caja patio sur 6 abril.xlsx".

Hoja clave: "FC X Obras" (Flujo de Caja por Obras)
Estructura:
- Filas: categorías agrupadas por tipo (Materiales, Mano Obra, Admon, Ingreso)
- Columnas: 39 meses (Oct 2025 → Dic 2028)
- Cada celda: valor mensual (proyectado o real)

Política: REEMPLAZA los valores existentes del proyecto.
"""
from __future__ import annotations
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


# Detección de grupos por keyword en encabezado de fila
# Soporta PATIO SUR y LYRA
GROUP_KEYWORDS = {
    "materiales": [
        "material", "accesorio", "aparato", "cable", "dotacion", "ducto",
        "generico", "herramienta", "luminaria", "papeleria", "redes",
        "seguridad", "servicio", "subestacion", "tablero", "tuberia",
        "voz y datos", "equipo", "feeder", "conduit", "conduit", "grounding",
        "pvc", "emt", "emt", "cable tray", "panel", "breaker", "disconnects",
        "excavacion", "conexion",
    ],
    "mano_obra": [
        "mano de obra", "mano obra", "personal operativ", "operativa",
        "tecnico", "ingeniero residente", "ingeniero de obra", "labor",
    ],
    "administracion": [
        "administracion", "administrativo", "directivo", "gerencia",
        "contabilidad", "seguridad social", "vehiculo", "oficina",
    ],
    "ingreso": [
        "ingreso", "facturacion", "anticipo", "factura", "cobro",
    ],
}

# Meses esperados (Oct 2025 → Dic 2028)
EXPECTED_MONTHS = [
    "2025-10", "2025-11", "2025-12",
    "2026-01", "2026-02", "2026-03", "2026-04", "2026-05", "2026-06",
    "2026-07", "2026-08", "2026-09", "2026-10", "2026-11", "2026-12",
    "2027-01", "2027-02", "2027-03", "2027-04", "2027-05", "2027-06",
    "2027-07", "2027-08", "2027-09", "2027-10", "2027-11", "2027-12",
    "2028-01", "2028-02", "2028-03", "2028-04", "2028-05", "2028-06",
    "2028-07", "2028-08", "2028-09", "2028-10", "2028-11", "2028-12",
]


def _normalize(text: Any) -> str:
    if not text:
        return ""
    s = str(text).strip().lower()
    repl = {"á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ñ": "n"}
    for k, v in repl.items():
        s = s.replace(k, v)
    return s


def _to_number(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).replace(",", "").replace("$", "").strip()
        if not s or s in ("-", "—"):
            return 0.0
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _detect_group(name: str) -> Optional[str]:
    """Detecta a qué grupo pertenece una fila por su nombre."""
    n = _normalize(name)
    for grupo, keywords in GROUP_KEYWORDS.items():
        for kw in keywords:
            if kw in n:
                return grupo
    return None


def _parse_month_header(cell_value: Any) -> Optional[str]:
    """Convierte encabezado de mes a formato YYYY-MM."""
    if not cell_value:
        return None
    if isinstance(cell_value, datetime):
        return f"{cell_value.year}-{cell_value.month:02d}"
    s = str(cell_value).strip()
    # Patrones tipo "Oct-25", "10/2025", "2025-10"
    m = re.search(r"(20\d{2})[\s\-/]?(\d{1,2})", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    m = re.search(r"(\d{1,2})[\s\-/]?(20\d{2})", s)
    if m:
        return f"{m.group(2)}-{int(m.group(1)):02d}"
    # Patrón "Oct-25" o "Ene 2025"
    months_es = {
        "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
        "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12,
    }
    m = re.search(r"([a-z]{3})[\s\-/]?(\d{2,4})", s.lower())
    if m and m.group(1)[:3] in months_es:
        year = int(m.group(2))
        if year < 100:
            year = 2000 + year
        return f"{year}-{months_es[m.group(1)[:3]]:02d}"
    return None


def _parse_relative_month_header(cell_value: Any) -> Optional[int]:
    """Convierte encabezado de mes relativo (Mes 1, Mes 2, ...) a número de mes."""
    if not cell_value:
        return None
    s = str(cell_value).strip()
    # Patrón "Mes 1", "Mes 2", ..., "Mes 18"
    m = re.search(r"Mes\s*(\d+)", s, re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None


def _calculate_mes_key_from_project_start(
    project_start_date: any, month_number: int
) -> str:
    """
    Calcula YYYY-MM basado en project_start_date + month_number relativo.
    month_number=1 → project_start_date month
    month_number=2 → project_start_date + 1 month

    Soporta: datetime object o string ISO (YYYY-MM-DD)
    """
    if not project_start_date:
        return ""

    # Si es string, convertir a datetime
    if isinstance(project_start_date, str):
        try:
            project_start_date = datetime.fromisoformat(project_start_date.replace('Z', '+00:00').split('T')[0] + 'T00:00:00')
        except:
            return ""

    start_year = project_start_date.year
    start_month = project_start_date.month

    # Calcular mes y año target
    total_months = start_month + (month_number - 1)
    target_year = start_year + (total_months - 1) // 12
    target_month = ((total_months - 1) % 12) + 1

    return f"{target_year}-{target_month:02d}"


def parse_cash_flow_excel(
    filepath: Path,
    sheet_hint: str = "FC X Obras",
    project_start_date: Optional[any] = None,  # datetime o string ISO (YYYY-MM-DD)
) -> Dict[str, Any]:
    """
    Parsea el Excel del Flujo de Caja. Soporta dos formatos:
    1. Meses calendario: "Oct-25", "10/2025", "2025-10"
    2. Meses relativos: "Mes 1", "Mes 2", ..., "Mes 18" (requiere project_start_date)

    Args:
        filepath: Ruta al archivo Excel
        sheet_hint: Nombre sugerido de la hoja
        project_start_date: Fecha de inicio del proyecto (datetime o string ISO YYYY-MM-DD)

    Returns:
        {
            "categorias": [
                {
                    "nombre": "ACCESORIOS GENERALES",
                    "grupo": "materiales",
                    "valores": {"2025-10": 0, "2026-02": 63396800, ...},
                    "sort_order": 1,
                    "row_index": 5,
                },
                ...
            ],
            "metadata": {
                "filename": "...",
                "sheet": "FC X Obras",
                "months_detected": [...],
                "total_categorias": 39,
                "total_valores": 1521,  # categorias * meses con valor
                "sum_total": 41012884481,
            }
        }
    """
    if not filepath.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {filepath}")

    # Convertir string a datetime si es necesario
    if isinstance(project_start_date, str):
        try:
            project_start_date = datetime.fromisoformat(
                project_start_date.replace('Z', '+00:00').split('T')[0] + 'T00:00:00'
            )
        except Exception:
            project_start_date = None

    wb = load_workbook(filepath, data_only=True, read_only=True)

    # Buscar hoja "FC X Obras", "FC", o similar
    sheet_name = None
    for s in wb.sheetnames:
        n = _normalize(s)
        if "fc x obras" in n or "fc_x_obras" in n or "flujo" in n:
            sheet_name = s
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]  # fallback al primero

    ws = wb[sheet_name]

    # Buscar fila de headers (la que tiene meses calendario O meses relativos)
    header_row_idx = None
    month_columns: Dict[int, str] = {}  # {col_index: "YYYY-MM"}
    has_relative_months = False

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        month_count = 0
        candidate = {}
        candidate_relative = {}

        for col_idx, cell in enumerate(row):
            # Intentar mes calendario primero
            mes = _parse_month_header(cell)
            if mes:
                candidate[col_idx] = mes
                month_count += 1
            else:
                # Intentar mes relativo
                month_num = _parse_relative_month_header(cell)
                if month_num:
                    candidate_relative[col_idx] = month_num
                    month_count += 1

        if month_count >= 5:  # al menos 5 columnas con meses → es la fila header
            header_row_idx = row_idx

            # Si tenemos meses relativos y start_date, convertir
            if candidate_relative and project_start_date:
                has_relative_months = True
                for col_idx, month_num in candidate_relative.items():
                    mes_key = _calculate_mes_key_from_project_start(project_start_date, month_num)
                    if mes_key:
                        month_columns[col_idx] = mes_key
            # Si tenemos meses calendario, usarlos
            elif candidate:
                month_columns = candidate

            break

    if not header_row_idx:
        wb.close()
        return {
            "categorias": [],
            "metadata": {
                "filename": filepath.name,
                "sheet": sheet_name,
                "error": "No se detectó fila de meses (calendario o relativo)",
            },
        }

    if has_relative_months and not project_start_date:
        wb.close()
        return {
            "categorias": [],
            "metadata": {
                "filename": filepath.name,
                "sheet": sheet_name,
                "error": "Se detectaron meses relativos pero no se proporcionó project_start_date",
            },
        }

    # Procesar filas debajo del header
    categorias = []
    sort_order = 0
    total_valores = 0
    sum_total = 0.0

    current_group: Optional[str] = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx <= header_row_idx:
            continue
        if not row:
            continue

        # Buscar nombre en las primeras 3 columnas
        name = None
        for c in row[:3]:
            if c and isinstance(c, str) and c.strip():
                name = c.strip()
                break
        if not name:
            continue

        # Detectar si es header de grupo
        # Soporta:
        # 1. PATIO SUR: texto en MAYÚSCULAS sin valores (o 1 valor)
        # 2. LYRA: nombres de grupo como "Feeders", "Excavacion", etc., SIN backslash
        #    (filas con backslash como "Feeder\8sets\..." son detail rows)
        nums_in_row = [v for v in row if isinstance(v, (int, float)) and v != 0]
        is_uppercase_header = name.isupper() and len(nums_in_row) <= 1

        grp = _detect_group(name)
        # LYRA group header: tiene keyword, no tiene backslash, pocas valores
        is_lyra_group_header = (
            grp and len(nums_in_row) <= 1 and "\\" not in name
        )

        if is_uppercase_header or is_lyra_group_header:
            if grp:
                current_group = grp
            continue

        # Determinar grupo: usar current_group o detectar por nombre
        grupo = _detect_group(name) or current_group
        if not grupo:
            continue  # ignorar filas sin grupo identificable

        # Extraer valores mensuales
        valores = {}
        for col_idx, mes in month_columns.items():
            if col_idx < len(row):
                val = _to_number(row[col_idx])
                if val != 0:
                    valores[mes] = val
                    total_valores += 1
                    sum_total += val

        if not valores:
            continue  # ignorar filas sin valores

        sort_order += 1
        categorias.append({
            "nombre": name,
            "grupo": grupo,
            "valores": valores,
            "sort_order": sort_order,
            "row_index": row_idx,
            "incluir_en_grafico": True,
        })

    wb.close()

    return {
        "categorias": categorias,
        "metadata": {
            "filename": filepath.name,
            "sheet": sheet_name,
            "months_detected": sorted(list(set(month_columns.values()))),
            "total_categorias": len(categorias),
            "total_valores": total_valores,
            "sum_total": sum_total,
            "has_relative_months": has_relative_months,
            "project_start_date": project_start_date.isoformat() if project_start_date else None,
            "parsed_at": datetime.now(timezone.utc).isoformat(),
        },
    }
