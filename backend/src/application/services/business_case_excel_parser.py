"""
Parser para el Excel "Detallado caso de negocio_220126.xlsx".

Lee las hojas clave y extrae los datos para poblar las tablas:
- Costo vs Venta            → business_case_chapters + business_case_aiu
- Ejecución vs Caso de Negocio → business_case_procurement + items
- Admon Patios              → business_case_indirect_costs
- RESUMEN VENTA             → business_case (KPIs)

Política: REEMPLAZA datos existentes del proyecto.
"""
from __future__ import annotations
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


# ──────────────────────────────────────────────────────────────────────────
# Mapeo de capítulos (chapter_id → label estándar usado en BD)
# ──────────────────────────────────────────────────────────────────────────

CHAPTER_NAME_MAP = {
    "redes mt": ("suministro", "Suministro", "redes-mt", "Redes MT (Celdas)"),
    "subestaciones": ("suministro", "Suministro", "subestaciones", "Subestaciones (Shelter)"),
    "transformadores": ("suministro", "Suministro", "transformadores", "Transformadores"),
    "baja tension": ("suministro", "Suministro", "bt", "Baja Tensión (BT)"),
    "spe y spt": ("suministro", "Suministro", "spe", "SPE y SPT"),
    "comunicaciones": ("suministro", "Suministro", "comunicaciones", "Comunicaciones"),
    "cargadores": ("suministro", "Suministro", "cargadores", "Cargadores"),
    "deteccion": ("suministro", "Suministro", "deteccion", "Detección Incendios"),
    "obras civiles": ("suministro", "Suministro", "obras-civiles", "Obras Civiles y Redes"),
    "estudios": ("mano-obra", "Mano de Obra", "estudios", "Estudios y Diseños"),
    "conexion": ("mano-obra", "Mano de Obra", "conexion-red", "Conexión a la Red"),
    "instalacion cargadores": ("mano-obra", "Mano de Obra", "inst-cargadores", "Instalación Cargadores"),
    "iluminacion": ("mano-obra", "Mano de Obra", "iluminacion", "ILU y Servicios Aux"),
    "tramites": ("administracion", "Administración", "tramites", "Trámites y Certificaciones"),
    "compensacion reactiva": ("intereses", "Intereses", "comp-reactiva", "Compensación Reactiva"),
}

AIU_KEYWORDS = {
    "iva cargadores": ("iva-cargadores", "IVA Cargadores", None),
    "its": ("its", "ITS", None),
    "administracion": ("adm-11", "Administración (11%)", 11.0),
    "imprevistos": ("imprev-2", "Imprevistos (2%)", 2.0),
    "utilidad": ("utilidad-4", "Utilidad (4%)", 4.0),
    "iva sobre": ("ivau-19", "IVA sobre Utilidad (19%)", 19.0),
    "financiacion": ("financiacion", "Financiación 9 m.", None),
}


def _normalize(text: Any) -> str:
    """Normaliza texto: minúsculas, sin tildes, sin espacios extras."""
    if not text:
        return ""
    s = str(text).strip().lower()
    replacements = {"á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ñ": "n"}
    for k, v in replacements.items():
        s = s.replace(k, v)
    return s


def _to_number(v: Any) -> float:
    """Convierte valor de Excel a float, retorna 0 si no es posible."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).replace(",", "").replace("$", "").strip()
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _match_chapter(name: str) -> Optional[Tuple[str, str, str, str]]:
    """Identifica chapter por nombre."""
    n = _normalize(name)
    for key, value in CHAPTER_NAME_MAP.items():
        if key in n:
            return value
    return None


def _match_aiu(name: str) -> Optional[Tuple[str, str, Optional[float]]]:
    """Identifica AIU item por keyword."""
    n = _normalize(name)
    for key, value in AIU_KEYWORDS.items():
        if key in n:
            return value
    return None


# ──────────────────────────────────────────────────────────────────────────
# Parsers por hoja
# ──────────────────────────────────────────────────────────────────────────

def parse_costo_vs_venta(filepath: Path) -> Dict[str, List[Dict]]:
    """
    Lee la hoja 'Costo vs Venta' y separa en chapters + AIU items.

    Estructura esperada del Excel (cols A-I):
      A: # ítem | B: Capítulo | ... | C: Venta | D: Costo

    Retorna:
        { "chapters": [...], "aiu": [...] }
    """
    wb = load_workbook(filepath, data_only=True, read_only=True)
    sheet_name = next((s for s in wb.sheetnames if "costo vs venta" in _normalize(s)), None)
    if not sheet_name:
        wb.close()
        return {"chapters": [], "aiu": []}

    ws = wb[sheet_name]
    chapters = []
    aiu_items = []
    chapter_order = 0
    aiu_order = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row or len(row) < 4:
            continue
        # Buscar nombre en col B y valores en cols posteriores
        name = row[1] if len(row) > 1 else None
        if not name:
            continue
        name_str = str(name).strip()

        # Buscar columnas con valores numéricos para venta/costo
        # Tomamos los últimos dos valores numéricos de la fila como (venta, costo)
        nums = [(i, _to_number(v)) for i, v in enumerate(row) if isinstance(v, (int, float))]
        if len(nums) < 2:
            continue
        venta = nums[-2][1] if len(nums) >= 2 else 0
        costo = nums[-1][1] if len(nums) >= 1 else 0

        # Match con chapter conocido
        ch_match = _match_chapter(name_str)
        if ch_match:
            grp_id, grp_name, ch_id, ch_name = ch_match
            chapter_order += 1
            chapters.append({
                "group_id": grp_id, "group_name": grp_name,
                "chapter_id": ch_id, "chapter_name": ch_name,
                "venta": venta, "costo": costo,
                "display_order": chapter_order,
                "excel_source_sheet": sheet_name,
                "excel_source_row": row_idx,
            })
            continue

        # Match con AIU
        aiu_match = _match_aiu(name_str)
        if aiu_match:
            tipo, label, pct = aiu_match
            aiu_order += 1
            aiu_items.append({
                "tipo": tipo, "label": label,
                "venta": venta, "costo": costo,
                "percentage": pct,
                "display_order": aiu_order,
            })

    wb.close()
    return {"chapters": chapters, "aiu": aiu_items}


def parse_resumen_venta(filepath: Path) -> Dict[str, Any]:
    """
    Lee la hoja 'RESUMEN VENTA' y extrae KPIs.

    Busca:
    - Valor Oferta Total
    - Costo Total
    - Margen Bruto
    - Administración
    - Financiación
    """
    wb = load_workbook(filepath, data_only=True, read_only=True)
    sheet_name = next((s for s in wb.sheetnames if "resumen venta" in _normalize(s)), None)
    if not sheet_name:
        wb.close()
        return {}

    ws = wb[sheet_name]
    kpis = {
        "valor_oferta_total": 0,
        "costo_total_sin_fin": 0,
        "costo_total_con_fin": 0,
        "margen_bruto_valor": 0,
        "margen_bruto_pct": 0,
        "administracion_valor": 0,
        "financiacion_valor": 0,
    }

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        text = " ".join(str(c) for c in row if c is not None).lower()
        nums = [_to_number(v) for v in row if isinstance(v, (int, float))]

        if "oferta" in text and "total" in text and nums:
            # En RESUMEN VENTA, el valor oferta suele ser el primer número grande
            kpis["valor_oferta_total"] = nums[0]
        elif "costo total" in text and "fin" in text and nums:
            if "sin" in text:
                kpis["costo_total_sin_fin"] = nums[-1]
            elif "con" in text:
                kpis["costo_total_con_fin"] = nums[-1]
        elif "margen" in text and nums:
            # Margen bruto en valor suele ser el último número grande
            kpis["margen_bruto_valor"] = nums[-1]
            pcts = [n for n in nums if 0 < n < 100]
            if pcts:
                kpis["margen_bruto_pct"] = max(pcts)
        elif "administracion" in text and nums:
            # Administración suele ser el último número (Costo)
            kpis["administracion_valor"] = nums[-1]
        elif "financiacion" in text and nums:
            # Financiación: el último número es el costo real (intereses)
            kpis["financiacion_valor"] = nums[-1]

    wb.close()
    return kpis


def parse_ejecucion_vs_cn(filepath: Path) -> List[Dict]:
    """
    Lee la hoja 'Ejecución vs Caso de Negocio' y extrae procurement.

    Estructura: cada bloque tiene un encabezado en MAYÚSCULAS y luego items.
    """
    wb = load_workbook(filepath, data_only=True, read_only=True)
    sheet_name = next(
        (s for s in wb.sheetnames if "ejecucion" in _normalize(s) and "caso" in _normalize(s)),
        None
    )
    if not sheet_name:
        wb.close()
        return []

    ws = wb[sheet_name]
    procurement = []
    current_group = None
    order = 0

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        first = row[0] if len(row) > 0 else None
        if not first:
            continue
        text = str(first).strip()

        # Detectar encabezado de capítulo (texto en mayúsculas)
        if text.isupper() and len(text) > 5 and len(text) < 60:
            order += 1
            nums = [_to_number(v) for v in row if isinstance(v, (int, float))]
            current_group = {
                "ref": text,
                "caso_negocio": nums[0] if len(nums) > 0 else 0,
                "negociado": nums[1] if len(nums) > 1 else 0,
                "pendiente": nums[2] if len(nums) > 2 else 0,
                "proyectado": nums[3] if len(nums) > 3 else 0,
                "display_order": order,
                "items": [],
            }
            procurement.append(current_group)

    wb.close()
    return procurement


# ──────────────────────────────────────────────────────────────────────────
# Función principal
# ──────────────────────────────────────────────────────────────────────────

def parse_business_case_excel(filepath: Path) -> Dict[str, Any]:
    """
    Punto de entrada: parsea el Excel completo y retorna estructura lista
    para insertar en BD.

    Returns:
        {
            "kpis": {...},
            "chapters": [...],
            "aiu": [...],
            "procurement": [...],
            "indirect_costs": [...],
        }
    """
    if not filepath.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {filepath}")

    cv = parse_costo_vs_venta(filepath)
    kpis = parse_resumen_venta(filepath)
    procurement = parse_ejecucion_vs_cn(filepath)

    return {
        "kpis": kpis,
        "chapters": cv["chapters"],
        "aiu": cv["aiu"],
        "procurement": procurement,
        "indirect_costs": [],  # TODO: parsear Admon Patios (256 cols, requiere análisis)
        "metadata": {
            "filename": filepath.name,
            "parsed_at": datetime.now(timezone.utc).isoformat(),
        },
    }
