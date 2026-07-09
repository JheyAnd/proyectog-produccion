"""
Endpoints REST para Caso de Negocio.

Política:
- Importar Excel → REEMPLAZA datos existentes
- Edición manual → registrada en business_case_audit_log
- Lecturas: GET completo o por sección
- Cambio de escenario: actualiza is_active

Endpoints:
  GET    /projects/{project_id}/business-case                    → Resumen + KPIs
  GET    /projects/{project_id}/business-case/full               → Todo en un payload
  GET    /projects/{project_id}/business-case/chapters           → Costo vs Venta
  GET    /projects/{project_id}/business-case/aiu                → Items AIU
  GET    /projects/{project_id}/business-case/procurement        → Detalle Costo
  GET    /projects/{project_id}/business-case/indirect-costs     → Costos Indirectos
  GET    /projects/{project_id}/business-case/scenarios          → Escenarios

  PUT    /projects/{project_id}/business-case/chapters/{id}      → Editar (con auditoría)
  PUT    /projects/{project_id}/business-case/aiu/{id}           → Editar
  POST   /projects/{project_id}/business-case/scenarios/activate → Cambiar escenario activo

  GET    /projects/{project_id}/business-case/audit-log          → Historial de ediciones
"""
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Optional, Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile, status
from pydantic import BaseModel
from sqlalchemy import select, and_, delete
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload



from src.application.services.business_case_excel_parser import parse_business_case_excel
from src.application.services.generic_budget_parser import detectar_total_venta, detectar_total_costo

from src.infrastructure.database.session import get_db_session
from src.infrastructure.database.models.business_case_model import (
    BusinessCaseModel,
    BusinessCaseChapterModel,
    BusinessCaseAIUModel,
    BusinessCaseProcurementModel,
    BusinessCaseProcurementItemModel,
    BusinessCaseIndirectCostModel,
    BusinessCaseScenarioModel,
    BusinessCaseAuditLogModel,
    BusinessCaseDetailModel,
)
from src.infrastructure.database.models.activity_model import ActivityLogModel
from src.infrastructure.database.models.user_model import UserModel
from src.infrastructure.database.models.project_file_model import ProjectFileModel
from src.interface.api.v1.dependencies.auth import get_current_user

import openpyxl
from io import BytesIO

router = APIRouter()


# ══════════════════════════════════════════════════════════════════════════════
# SCHEMAS
# ══════════════════════════════════════════════════════════════════════════════

class BusinessCaseResponse(BaseModel):
    id: str
    project_id: str
    scenario_active: str
    usd_rate: float
    valor_oferta_total: float
    costo_total_sin_fin: float
    costo_total_con_fin: float
    margen_bruto_valor: float
    margen_bruto_pct: float
    administracion_valor: float
    financiacion_valor: float
    meses_sin_ingresos: Optional[int] = None
    source_excel_filename: Optional[str] = None
    last_imported_at: Optional[datetime] = None
    last_imported_by_name: Optional[str] = None
    updated_at: datetime
    venta_monto_manual: Optional[float] = None
    venta_materiales: Optional[float] = None
    venta_servicios: Optional[float] = None
    venta_administracion: Optional[float] = None
    venta_mano_obra: Optional[float] = None
    venta_intereses: Optional[float] = None
    costo_monto_manual: Optional[float] = None
    costo_materiales: Optional[float] = None
    costo_servicios: Optional[float] = None
    costo_administracion: Optional[float] = None
    costo_mano_obra: Optional[float] = None
    costo_intereses: Optional[float] = None
    valores_manuales_completos: Optional[bool] = False
    venta_excel_validado: Optional[bool] = False
    costo_excel_validado: Optional[bool] = False
    model_config = {"from_attributes": True}


class ChapterResponse(BaseModel):
    id: str
    business_case_id: str
    group_id: str
    group_name: str
    chapter_id: str
    chapter_name: str
    venta: float
    costo: float
    display_order: int
    model_config = {"from_attributes": True}


class AIUResponse(BaseModel):
    id: str
    business_case_id: str
    tipo: str
    label: str
    venta: float
    costo: float
    percentage: Optional[float] = None
    display_order: int
    model_config = {"from_attributes": True}


class ProcurementItemResponse(BaseModel):
    id: str
    capitulo: str
    proveedor: Optional[str] = None
    negociado: float
    pendiente: float
    display_order: int
    model_config = {"from_attributes": True}


class ProcurementResponse(BaseModel):
    id: str
    ref: str
    caso_negocio: float
    negociado: float
    pendiente: float
    proyectado: float
    display_order: int
    items: List[ProcurementItemResponse] = []
    model_config = {"from_attributes": True}


class IndirectCostResponse(BaseModel):
    id: str
    seccion: str
    item_code: Optional[str] = None
    descripcion: str
    unidad: Optional[str] = None
    cantidad: Optional[float] = None
    vr_unitario: Optional[float] = None
    total: float
    display_order: int
    model_config = {"from_attributes": True}


class ScenarioResponse(BaseModel):
    id: str
    scenario_name: str
    usd_rate: float
    total_oferta: float
    total_costo: float
    margen_pct: float
    is_active: bool
    model_config = {"from_attributes": True}


class AuditLogResponse(BaseModel):
    id: str
    table_name: str
    record_id: str
    field_name: str
    old_value: Optional[str] = None
    new_value: Optional[str] = None
    user_name: str
    user_role: str
    occurred_at: datetime
    notes: Optional[str] = None
    model_config = {"from_attributes": True}


class BusinessCaseDetailResponse(BaseModel):
    id: str
    project_id: str
    tipo: str
    categoria: str
    concepto: str
    valor: float
    moneda: str
    fuente_excel: Optional[str] = None
    creado_por_ia: int
    created_at: datetime
    model_config = {"from_attributes": True}


class ParseDetailRequest(BaseModel):
    tipo: str


class FullBusinessCaseResponse(BaseModel):
    """Payload completo para el frontend en una sola llamada."""
    business_case: BusinessCaseResponse
    chapters: List[ChapterResponse]
    aiu: List[AIUResponse]
    procurement: List[ProcurementResponse]
    indirect_costs: List[IndirectCostResponse]
    scenarios: List[ScenarioResponse]


# Edit schemas
class ChapterUpdateRequest(BaseModel):
    venta: Optional[float] = None
    costo: Optional[float] = None
    notes: Optional[str] = None


class AIUUpdateRequest(BaseModel):
    venta: Optional[float] = None
    costo: Optional[float] = None
    notes: Optional[str] = None


class ScenarioActivateRequest(BaseModel):
    scenario_id: str


class ValoresManualesRequest(BaseModel):
    venta_monto_manual: float
    venta_materiales: float
    venta_servicios: float
    venta_administracion: float
    venta_mano_obra: float
    venta_intereses: Optional[float] = 0.0
    costo_monto_manual: float
    costo_materiales: float
    costo_servicios: float
    costo_administracion: float
    costo_mano_obra: float
    costo_intereses: Optional[float] = 0.0
    trm: Optional[float] = None


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

CARSAN_COMPANY_ID = 3

async def es_proyecto_usd(project_id: str, db: AsyncSession) -> bool:
    """Retorna True si el proyecto pertenece a CARSAN (USD)"""
    import sqlalchemy
    proyecto = await db.execute(
        sqlalchemy.text("SELECT company_id FROM projects WHERE id = :pid"),
        {"pid": project_id}
    )
    res = proyecto.fetchone()
    if not res:
        # Buscar en project_tracking
        proyecto = await db.execute(
            sqlalchemy.text("SELECT company_id FROM project_tracking WHERE id = :pid OR project_id = :pid"),
            {"pid": project_id}
        )
        res = proyecto.fetchone()
    return bool(res and res.company_id == CARSAN_COMPANY_ID)

def _get_client_ip(request: Request) -> Optional[str]:
    fwd = request.headers.get("X-Forwarded-For")
    if fwd:
        return fwd.split(",")[0].strip()
    return request.client.host if request.client else None


async def _log_edit(
    db: AsyncSession,
    business_case_id: str,
    project_id: str,
    table_name: str,
    record_id: str,
    field_name: str,
    old_value: Any,
    new_value: Any,
    user: UserModel,
    request: Request,
    notes: Optional[str] = None,
):
    """Registra una edición en business_case_audit_log."""
    db.add(BusinessCaseAuditLogModel(
        id=str(uuid.uuid4()),
        business_case_id=business_case_id,
        project_id=project_id,
        table_name=table_name,
        record_id=record_id,
        field_name=field_name,
        old_value=str(old_value) if old_value is not None else None,
        new_value=str(new_value) if new_value is not None else None,
        user_id=user.id,
        user_name=user.full_name,
        user_role=user.role,
        ip_address=_get_client_ip(request),
        notes=notes,
    ))
    
    # También registrar en la tabla global activity_logs para visibilidad centralizada
    db.add(ActivityLogModel(
        user_id=str(user.id),
        user_name=user.full_name or user.email,
        user_role=user.role,
        module="business_case",
        page=f"projects/{project_id}/business-case",
        action=f"update_bc_{table_name}: {field_name}",
        field_name=field_name,
        before_state=str(old_value) if old_value is not None else None,
        after_state=str(new_value) if new_value is not None else None,
        target_link=f"/projects/{project_id}/business-case",
        project_id=project_id,
        timestamp=datetime.now(timezone.utc),
    ))


async def _get_business_case(db: AsyncSession, project_id: str) -> BusinessCaseModel:
    result = await db.execute(
        select(BusinessCaseModel).where(BusinessCaseModel.project_id == project_id)
    )
    bc = result.scalar_one_or_none()
    if not bc:
        raise HTTPException(status_code=404, detail=f"Caso de Negocio no existe para proyecto {project_id}")
    return bc


# ══════════════════════════════════════════════════════════════════════════════
# READ ENDPOINTS
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/projects/{project_id}/business-case", response_model=BusinessCaseResponse)
async def get_summary(project_id: str, db: AsyncSession = Depends(get_db_session)):
    """KPIs principales del Caso de Negocio (los 4 cards superiores)."""
    bc = await _get_business_case(db, project_id)
    return bc


@router.get("/projects/{project_id}/business-case/full", response_model=FullBusinessCaseResponse)
async def get_full(project_id: str, db: AsyncSession = Depends(get_db_session)):
    """Payload COMPLETO en una sola llamada (para optimizar carga inicial)."""
    bc = await _get_business_case(db, project_id)

    # Chapters
    ch_result = await db.execute(
        select(BusinessCaseChapterModel)
        .where(BusinessCaseChapterModel.project_id == project_id)
        .order_by(BusinessCaseChapterModel.display_order)
    )
    chapters = list(ch_result.scalars().all())

    # AIU
    aiu_result = await db.execute(
        select(BusinessCaseAIUModel)
        .where(BusinessCaseAIUModel.project_id == project_id)
        .order_by(BusinessCaseAIUModel.display_order)
    )
    aiu = list(aiu_result.scalars().all())

    # Procurement con items cargados vía selectinload para evitar MissingGreenlet en Pydantic
    proc_result = await db.execute(
        select(BusinessCaseProcurementModel)
        .where(BusinessCaseProcurementModel.project_id == project_id)
        .options(selectinload(BusinessCaseProcurementModel.items))
        .order_by(BusinessCaseProcurementModel.display_order)
    )
    procs = list(proc_result.scalars().all())

    proc_responses = [ProcurementResponse.model_validate(p) for p in procs]

    # Indirect costs
    ic_result = await db.execute(
        select(BusinessCaseIndirectCostModel)
        .where(BusinessCaseIndirectCostModel.project_id == project_id)
        .order_by(BusinessCaseIndirectCostModel.display_order)
    )
    indirect = list(ic_result.scalars().all())

    # Scenarios
    sc_result = await db.execute(
        select(BusinessCaseScenarioModel)
        .where(BusinessCaseScenarioModel.project_id == project_id)
        .order_by(BusinessCaseScenarioModel.usd_rate)
    )
    scenarios = list(sc_result.scalars().all())

    return FullBusinessCaseResponse(
        business_case=BusinessCaseResponse.model_validate(bc),
        chapters=[ChapterResponse.model_validate(c) for c in chapters],
        aiu=[AIUResponse.model_validate(a) for a in aiu],
        procurement=proc_responses,
        indirect_costs=[IndirectCostResponse.model_validate(i) for i in indirect],
        scenarios=[ScenarioResponse.model_validate(s) for s in scenarios],
    )


@router.get("/projects/{project_id}/business-case/chapters", response_model=List[ChapterResponse])
async def get_chapters(project_id: str, db: AsyncSession = Depends(get_db_session)):
    """Capítulos de Costo vs Venta."""
    bc = await _get_business_case(db, project_id)
    result = await db.execute(
        select(BusinessCaseChapterModel)
        .where(BusinessCaseChapterModel.project_id == project_id)
        .order_by(BusinessCaseChapterModel.display_order)
    )
    return list(result.scalars().all())


@router.get("/projects/{project_id}/business-case/aiu", response_model=List[AIUResponse])
async def get_aiu(project_id: str, db: AsyncSession = Depends(get_db_session)):
    """Items AIU + Financiación."""
    bc = await _get_business_case(db, project_id)
    result = await db.execute(
        select(BusinessCaseAIUModel)
        .where(BusinessCaseAIUModel.project_id == project_id)
        .order_by(BusinessCaseAIUModel.display_order)
    )
    return list(result.scalars().all())


@router.get("/projects/{project_id}/business-case/procurement", response_model=List[ProcurementResponse])
async def get_procurement(project_id: str, db: AsyncSession = Depends(get_db_session)):
    # Detalle Costo con items cargados vía selectinload
    result = await db.execute(
        select(BusinessCaseProcurementModel)
        .where(BusinessCaseProcurementModel.project_id == project_id)
        .options(selectinload(BusinessCaseProcurementModel.items))
        .order_by(BusinessCaseProcurementModel.display_order)
    )
    procs = list(result.scalars().all())
    return [ProcurementResponse.model_validate(p) for p in procs]


@router.get("/projects/{project_id}/business-case/indirect-costs", response_model=List[IndirectCostResponse])
async def get_indirect_costs(project_id: str, db: AsyncSession = Depends(get_db_session)):
    """Costos Indirectos (Admon Patios)."""
    bc = await _get_business_case(db, project_id)
    result = await db.execute(
        select(BusinessCaseIndirectCostModel)
        .where(BusinessCaseIndirectCostModel.project_id == project_id)
        .order_by(BusinessCaseIndirectCostModel.display_order)
    )
    return list(result.scalars().all())


@router.get("/projects/{project_id}/business-case/scenarios", response_model=List[ScenarioResponse])
async def get_scenarios(project_id: str, db: AsyncSession = Depends(get_db_session)):
    """Escenarios disponibles (USD3900/4000/4300)."""
    bc = await _get_business_case(db, project_id)
    result = await db.execute(
        select(BusinessCaseScenarioModel)
        .where(BusinessCaseScenarioModel.project_id == project_id)
        .order_by(BusinessCaseScenarioModel.usd_rate)
    )
    return list(result.scalars().all())


@router.get("/projects/{project_id}/business-case/audit-log", response_model=List[AuditLogResponse])
async def get_audit_log(
    project_id: str,
    limit: int = 100,
    db: AsyncSession = Depends(get_db_session),
):
    """Historial de ediciones del Caso de Negocio."""
    bc = await _get_business_case(db, project_id)
    result = await db.execute(
        select(BusinessCaseAuditLogModel)
        .where(BusinessCaseAuditLogModel.project_id == project_id)
        .order_by(BusinessCaseAuditLogModel.occurred_at.desc())
        .limit(limit)
    )
    return list(result.scalars().all())


# ══════════════════════════════════════════════════════════════════════════════
# WRITE ENDPOINTS (Editable con auditoría)
# ══════════════════════════════════════════════════════════════════════════════

@router.put("/projects/{project_id}/business-case/chapters/{chapter_id}", response_model=ChapterResponse)
async def update_chapter(
    project_id: str,
    chapter_id: str,
    body: ChapterUpdateRequest,
    request: Request,
    current_user: UserModel = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session),
):
    """Edita un capítulo de Costo vs Venta. Cada cambio queda auditado."""
    # Patio Sur ya no está protegido
    bc = await _get_business_case(db, project_id)

    result = await db.execute(
        select(BusinessCaseChapterModel).where(
            and_(
                BusinessCaseChapterModel.id == chapter_id,
                BusinessCaseChapterModel.business_case_id == bc.id,
            )
        )
    )
    chapter = result.scalar_one_or_none()
    if not chapter:
        raise HTTPException(status_code=404, detail="Capítulo no encontrado")

    if body.venta is not None and body.venta != float(chapter.venta):
        await _log_edit(db, bc.id, project_id, "business_case_chapters", chapter.id,
                        "venta", chapter.venta, body.venta, current_user, request, body.notes)
        chapter.venta = body.venta

    if body.costo is not None and body.costo != float(chapter.costo):
        await _log_edit(db, bc.id, project_id, "business_case_chapters", chapter.id,
                        "costo", chapter.costo, body.costo, current_user, request, body.notes)
        chapter.costo = body.costo

    chapter.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(chapter)
    return chapter


@router.put("/projects/{project_id}/business-case/aiu/{aiu_id}", response_model=AIUResponse)
async def update_aiu(
    project_id: str,
    aiu_id: str,
    body: AIUUpdateRequest,
    request: Request,
    current_user: UserModel = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session),
):
    """Edita un item AIU. Cada cambio queda auditado."""
    # Patio Sur ya no está protegido
    bc = await _get_business_case(db, project_id)

    result = await db.execute(
        select(BusinessCaseAIUModel).where(
            and_(
                BusinessCaseAIUModel.id == aiu_id,
                BusinessCaseAIUModel.business_case_id == bc.id,
            )
        )
    )
    aiu = result.scalar_one_or_none()
    if not aiu:
        raise HTTPException(status_code=404, detail="Item AIU no encontrado")

    if body.venta is not None and body.venta != float(aiu.venta):
        await _log_edit(db, bc.id, project_id, "business_case_aiu", aiu.id,
                        "venta", aiu.venta, body.venta, current_user, request, body.notes)
        aiu.venta = body.venta

    if body.costo is not None and body.costo != float(aiu.costo):
        await _log_edit(db, bc.id, project_id, "business_case_aiu", aiu.id,
                        "costo", aiu.costo, body.costo, current_user, request, body.notes)
        aiu.costo = body.costo

    aiu.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(aiu)
    return aiu


@router.post("/projects/{project_id}/business-case/scenarios/activate", response_model=ScenarioResponse)
async def activate_scenario(
    project_id: str,
    body: ScenarioActivateRequest,
    request: Request,
    current_user: UserModel = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session),
):
    """Cambia el escenario activo (USD3900/4000/4300)."""
    # Patio Sur ya no está protegido
    bc = await _get_business_case(db, project_id)

    # Validar que el escenario existe
    result = await db.execute(
        select(BusinessCaseScenarioModel).where(
            and_(
                BusinessCaseScenarioModel.id == body.scenario_id,
                BusinessCaseScenarioModel.business_case_id == bc.id,
            )
        )
    )
    new_active = result.scalar_one_or_none()
    if not new_active:
        raise HTTPException(status_code=404, detail="Escenario no encontrado")

    # Desactivar todos los escenarios actuales
    all_result = await db.execute(
        select(BusinessCaseScenarioModel).where(
            BusinessCaseScenarioModel.business_case_id == bc.id
        )
    )
    for sc in all_result.scalars().all():
        sc.is_active = (sc.id == body.scenario_id)

    # Auditoría
    old_scenario = bc.scenario_active
    await _log_edit(db, bc.id, project_id, "business_case", bc.id,
                    "scenario_active", old_scenario, new_active.scenario_name,
                    current_user, request,
                    notes=f"Cambio de escenario USD")

    # Actualizar bc
    bc.scenario_active = new_active.scenario_name
    bc.usd_rate = new_active.usd_rate
    bc.valor_oferta_total = new_active.total_oferta
    bc.costo_total_con_fin = new_active.total_costo
    bc.margen_bruto_pct = new_active.margen_pct
    bc.updated_at = datetime.now(timezone.utc)

    await db.commit()
    await db.refresh(new_active)
    return new_active


# ══════════════════════════════════════════════════════════════════════════════
# IMPORTAR EXCEL (REEMPLAZA datos existentes)
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/projects/{project_id}/business-case/import-excel")
async def import_excel(
    project_id: str,
    request: Request,
    file: UploadFile = File(...),
    current_user: UserModel = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session),
):
    """
    Importa un Excel de Caso de Negocio. POLÍTICA: REEMPLAZA datos existentes.

    El archivo se guarda temporalmente, se parsea, y los datos se insertan
    eliminando los anteriores. Queda registrado quién importó cuándo.
    """
    # Solo Controller, Gerente o Administrador pueden importar
    # Patio Sur ya no está protegido
    
    if current_user.role not in ("controller", "gerente", "administrador"):
        raise HTTPException(status_code=403, detail="Sin permiso para importar Excel")

    # Validar extensión
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo archivos Excel (.xlsx, .xls)")

    # Guardar temporalmente
    import tempfile
    suffix = Path(file.filename).suffix
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = Path(tmp.name)

    try:
        # Parsear Excel
        parsed = parse_business_case_excel(tmp_path)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parseando Excel: {e}")
    finally:
        try:
            tmp_path.unlink()
        except Exception:
            pass

    # REEMPLAZAR datos existentes
    existing = await db.execute(
        select(BusinessCaseModel).where(BusinessCaseModel.project_id == project_id)
    )
    old_bc = existing.scalar_one_or_none()
    if old_bc:
        await db.execute(delete(BusinessCaseModel).where(BusinessCaseModel.id == old_bc.id))
        await db.flush()

    # Crear nuevo business_case
    bc_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    kpis = parsed.get("kpis", {})

    bc = BusinessCaseModel(
        id=bc_id,
        project_id=project_id,
        scenario_active="USD4000",
        usd_rate=4000.00,
        valor_oferta_total=kpis.get("valor_oferta_total", 0),
        costo_total_sin_fin=kpis.get("costo_total_sin_fin", 0),
        costo_total_con_fin=kpis.get("costo_total_con_fin", 0),
        margen_bruto_valor=kpis.get("margen_bruto_valor", 0),
        margen_bruto_pct=kpis.get("margen_bruto_pct", 0),
        administracion_valor=kpis.get("administracion_valor", 0),
        financiacion_valor=kpis.get("financiacion_valor", 0),
        source_excel_filename=file.filename,
        last_imported_at=now,
        last_imported_by_id=current_user.id,
        last_imported_by_name=current_user.full_name,
        created_at=now,
        updated_at=now,
    )
    db.add(bc)
    await db.flush()

    # Insertar chapters
    for ch in parsed.get("chapters", []):
        db.add(BusinessCaseChapterModel(
            id=str(uuid.uuid4()),
            business_case_id=bc_id,
            project_id=project_id,
            **ch,
        ))

    # Insertar AIU
    for a in parsed.get("aiu", []):
        db.add(BusinessCaseAIUModel(
            id=str(uuid.uuid4()),
            business_case_id=bc_id,
            project_id=project_id,
            **a,
        ))

    # Insertar procurement
    for p in parsed.get("procurement", []):
        db.add(BusinessCaseProcurementModel(
            id=str(uuid.uuid4()),
            business_case_id=bc_id,
            project_id=project_id,
            **p,
        ))

    # Insertar indirect costs
    for ic in parsed.get("indirect_costs", []):
        db.add(BusinessCaseIndirectCostModel(
            id=str(uuid.uuid4()),
            business_case_id=bc_id,
            project_id=project_id,
            **ic,
        ))

    # Insertar scenarios
    for sc in parsed.get("scenarios", []):
        db.add(BusinessCaseScenarioModel(
            id=str(uuid.uuid4()),
            business_case_id=bc_id,
            project_id=project_id,
            **sc,
        ))

    # Auditoría: registrar la importación como evento
    await _log_edit(
        db, bc_id, project_id, "business_case", bc_id, "_import_excel_",
        old_value=(old_bc.source_excel_filename if old_bc else None),
        new_value=file.filename,
        user=current_user, request=request,
        notes=f"Importación masiva: REEMPLAZÓ todos los datos del Caso de Negocio",
    )

    await db.commit()

    return {
        "status": "imported",
        "filename": file.filename,
        "imported_by": current_user.full_name,
        "imported_at": now.isoformat(),
        "summary": {
            "chapters": len(parsed.get("chapters", [])),
            "aiu_items": len(parsed.get("aiu", [])),
            "procurement": len(parsed.get("procurement", [])),
            "kpis": kpis,
        },
        "warning": "Se REEMPLAZARON todos los datos anteriores del Caso de Negocio.",
    }


@router.post("/projects/{project_id}/business-case/import/lyra-carsan")
async def import_lyra_carsan(
    project_id: str,
    file: UploadFile = File(...),
    trm: Optional[float] = None,
    usd_rate: Optional[float] = Form(None),
    tipo: Optional[str] = Form(None),
    current_user: UserModel = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session)
):
    """
    Endpoint especializado para importar presupuestos (venta o costo) de LYRA CARSAN.
    """
    from io import BytesIO
    import openpyxl

    # Patio Sur ya no está protegido

    # 1. Determinar el tipo (venta o costo)
    effective_tipo = tipo
    if not effective_tipo:
        fn = file.filename.lower()
        if "venta" in fn or "oferta" in fn:
            effective_tipo = "venta"
        elif "costo" in fn or "presupuesto" in fn:
            effective_tipo = "costo"
        else:
            raise HTTPException(status_code=400, detail="Debe especificar el tipo ('venta' o 'costo') en el formulario o en el nombre del archivo.")

    if effective_tipo not in ("venta", "costo"):
        raise HTTPException(status_code=400, detail=f"Tipo inválido: '{effective_tipo}'. Debe ser 'venta' o 'costo'.")

    # 2. Leer archivo y parsear
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error leyendo el archivo Excel: {str(e)}")

    if effective_tipo == "venta":
        try:
            total_venta = detectar_total_venta(wb)
        except ValueError as ve:
            raise HTTPException(status_code=400, detail=str(ve))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error al detectar total de venta: {str(e)}")
    else:
        try:
            total_costo = detectar_total_costo(wb)
        except ValueError as ve:
            raise HTTPException(status_code=400, detail=str(ve))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error al detectar total de costo: {str(e)}")

    # 3. Conversión de moneda si el proyecto es USD
    es_usd = await es_proyecto_usd(project_id, db)
    moneda_val = "USD" if es_usd else "COP"

    effective_trm = trm or usd_rate
    result_bc = await db.execute(
        select(BusinessCaseModel).where(BusinessCaseModel.project_id == project_id)
    )
    bc = result_bc.scalar_one_or_none()

    if not effective_trm and bc and bc.usd_rate:
        effective_trm = bc.usd_rate

    if not effective_trm:
        effective_trm = 4000.0

    if es_usd and effective_trm:
        if effective_tipo == "venta":
            total_venta = total_venta * effective_trm
        else:
            total_costo = total_costo * effective_trm

    # 4. Actualizar o crear BusinessCase
    if not bc:
        bc = BusinessCaseModel(
            id=str(uuid.uuid4()),
            project_id=project_id,
            valor_oferta_total=total_venta if effective_tipo == 'venta' else 0,
            costo_total=total_costo if effective_tipo == 'costo' else 0,
            costo_total_sin_fin=total_costo if effective_tipo == 'costo' else 0,
            costo_total_con_fin=total_costo if effective_tipo == 'costo' else 0,
            presupuesto_venta_cargado=(effective_tipo == 'venta'),
            presupuesto_costo_cargado=(effective_tipo == 'costo'),
            moneda=moneda_val
        )
        if es_usd and effective_trm:
            bc.usd_rate = effective_trm
        db.add(bc)
    else:
        if effective_tipo == 'venta':
            bc.valor_oferta_total = total_venta
            bc.presupuesto_venta_cargado = True
        else:
            bc.costo_total = total_costo
            bc.costo_total_sin_fin = total_costo
            bc.costo_total_con_fin = total_costo
            bc.presupuesto_costo_cargado = True
        bc.moneda = moneda_val
        if es_usd and effective_trm:
            bc.usd_rate = effective_trm
        bc.updated_at = datetime.now(timezone.utc)

    # 5. Guardar archivo en project_files
    new_file = ProjectFileModel(
        project_id=project_id,
        nombre_original=file.filename,
        tipo_mime=file.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        tamano_bytes=len(content),
        categoria="presupuesto_venta" if effective_tipo == 'venta' else "presupuesto_costo",
        archivo=content,
        subido_por=current_user.full_name or current_user.email
    )
    db.add(new_file)

    # 6. Registrar actividad
    db.add(ActivityLogModel(
        user_id=str(current_user.id),
        user_name=current_user.full_name or current_user.email,
        user_role=current_user.role,
        module="business_case",
        page=f"projects/{project_id}/business-case",
        action=f"upload_{effective_tipo}_lyra: {file.filename}",
        field_name="valor_oferta_total" if effective_tipo == 'venta' else "costo_total_sin_fin",
        after_state=str(total_venta if effective_tipo == 'venta' else total_costo),
        project_id=project_id,
        timestamp=datetime.now(timezone.utc),
    ))

    await db.commit()

    if effective_tipo == "venta":
        return {
            "total_venta": total_venta,
            "archivo": file.filename,
            "moneda": moneda_val,
            "trm": effective_trm
        }
    else:
        return {
            "total_costo": total_costo,
            "archivo": file.filename,
            "moneda": moneda_val,
            "trm": effective_trm
        }


# ══════════════════════════════════════════════════════════════════════════════
# CARGA DE ENTREGABLES (PRESUPUESTOS) — DETECCIÓN AUTOMÁTICA
# ══════════════════════════════════════════════════════════════════════════════

PROYECTOS_PROTEGIDOS = []
PROYECTOS_SIN_PARSER_GENERICO = ['lyra-carsan-oe2000']

@router.get("/projects/{project_id}/business-case/status")
async def get_business_case_status(
    project_id: str,
    db: AsyncSession = Depends(get_db_session),
    current_user: UserModel = Depends(get_current_user)
):
    """Retorna el estado de los entregables para desbloquear el Caso de Negocio."""
    try:
        result = await db.execute(
            select(BusinessCaseModel).where(BusinessCaseModel.project_id == project_id)
        )
        bc = result.scalar_one_or_none()
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print("EXCEPTION IN GET_STATUS:", error_details)
        return {"error": str(e), "traceback": error_details, "venta_cargado": False, "costo_cargado": False}
    
    if not bc:
        # Si no existe, retornamos valores por defecto robustos
        from src.infrastructure.database.models.project_model import ProjectModel
        proj_res = await db.execute(select(ProjectModel).where(ProjectModel.id == project_id))
        proj = proj_res.scalar_one_or_none()
        moneda_default = "USD" if (proj and (proj.currency == "USD" or proj.company_id == 3)) else "COP"
        return {
            "venta_cargado": False,
            "costo_cargado": False,
            "valor_oferta_total": 0,
            "costo_total": 0,
            "valores_manuales_completos": False,
            "venta_monto_manual": 0,
            "venta_materiales": 0,
            "venta_servicios": 0,
            "venta_administracion": 0,
            "venta_mano_obra": 0,
            "venta_intereses": 0,
            "costo_monto_manual": 0,
            "costo_materiales": 0,
            "costo_servicios": 0,
            "costo_administracion": 0,
            "costo_mano_obra": 0,
            "costo_intereses": 0,
            "venta_excel_validado": False,
            "costo_excel_validado": False,
            "moneda": moneda_default,
            "usd_rate": 4000
        }
    
    return {
        "venta_cargado": bc.presupuesto_venta_cargado,
        "costo_cargado": bc.presupuesto_costo_cargado,
        "valor_oferta_total": bc.valor_oferta_total,
        "costo_total": bc.costo_total,
        # Nuevos campos Paso 1
        "valores_manuales_completos": bc.valores_manuales_completos,
        "venta_monto_manual": bc.venta_monto_manual,
        "venta_materiales": bc.venta_materiales,
        "venta_servicios": bc.venta_servicios,
        "venta_administracion": bc.venta_administracion,
        "venta_mano_obra": bc.venta_mano_obra,
        "venta_intereses": bc.venta_intereses or 0,
        "costo_monto_manual": bc.costo_monto_manual,
        "costo_materiales": bc.costo_materiales,
        "costo_servicios": bc.costo_servicios,
        "costo_administracion": bc.costo_administracion,
        "costo_mano_obra": bc.costo_mano_obra,
        "costo_intereses": bc.costo_intereses or 0,
        # Flags de validación
        "venta_excel_validado": bc.venta_excel_validado,
        "costo_excel_validado": bc.costo_excel_validado,
        "moneda": bc.moneda,
        "usd_rate": bc.usd_rate
    }


@router.patch("/projects/{project_id}/business-case/valores-manuales")
async def guardar_valores_manuales(
    project_id: str,
    body: ValoresManualesRequest,
    current_user: UserModel = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session)
):
    """Guarda los valores manuales (Paso 1) antes de la carga de Excel."""
    from sqlalchemy import text
    
    # ── Protección Patio Sur ──────────────────────────────────────────
    if project_id == 'patio-sur-oe1035':
        raise HTTPException(status_code=403, detail="Patio Sur está protegido y no permite modificaciones.")

    # ── Verificar que el proyecto existe ─────────────────────────────
    res_proj = await db.execute(
        text("SELECT id FROM projects WHERE id = :pid"),
        {"pid": project_id}
    )
    proyecto = res_proj.fetchone()
    if not proyecto:
        raise HTTPException(status_code=404, detail=f"Proyecto '{project_id}' no encontrado.")

    # ── UPSERT: crear registro si no existe ──────────────────────────
    res_bc = await db.execute(
        text("SELECT id FROM business_case WHERE project_id = :pid"),
        {"pid": project_id}
    )
    existe = res_bc.fetchone()

    if not existe:
        await db.execute(
            text("""
                INSERT INTO business_case (
                    id, project_id, scenario_active, usd_rate, valor_oferta_total,
                    costo_total_sin_fin, costo_total_con_fin, margen_bruto_valor,
                    margen_bruto_pct, administracion_valor, financiacion_valor,
                    created_at, updated_at
                ) VALUES (
                    :id, :pid, 'USD4000', 4000.00, 0,
                    0, 0, 0,
                    0, 0, 0,
                    NOW(), NOW()
                )
            """),
            {"id": str(uuid.uuid4()), "pid": project_id}
        )
        await db.commit()

    # ── Guardar valores manuales ─────────────────────────────────────
    await db.execute(text("""
        UPDATE business_case SET
            venta_monto_manual         = :vm,
            venta_materiales           = :vmat,
            venta_servicios            = :vser,
            venta_administracion       = :va,
            venta_mano_obra            = :vmo,
            venta_intereses            = :vi,
            costo_monto_manual         = :cm,
            costo_materiales           = :cmat,
            costo_servicios            = :cser,
            costo_administracion       = :ca,
            costo_mano_obra            = :cmo,
            costo_intereses            = :ci,
            valores_manuales_completos = TRUE
        WHERE project_id = :pid
    """), {
        "vm":   body.venta_monto_manual,
        "vmat": body.venta_materiales,
        "vser": body.venta_servicios,
        "va":   body.venta_administracion,
        "vmo":  body.venta_mano_obra,
        "vi":   body.venta_intereses,
        "cm":   body.costo_monto_manual,
        "cmat": body.costo_materiales,
        "cser": body.costo_servicios,
        "ca":   body.costo_administracion,
        "cmo":  body.costo_mano_obra,
        "ci":   body.costo_intereses,
        "pid":  project_id
    })
    
    # Manejo de moneda
    es_usd = await es_proyecto_usd(project_id, db)
    moneda_val = "USD" if es_usd else "COP"
    await db.execute(
        text("UPDATE business_case SET moneda = :moneda WHERE project_id = :pid"),
        {"moneda": moneda_val, "pid": project_id}
    )
    
    if es_usd and body.trm:
        await db.execute(
            text("UPDATE business_case SET usd_rate = :trm WHERE project_id = :pid"),
            {"trm": body.trm, "pid": project_id}
        )

    # 2. Registrar actividad
    db.add(ActivityLogModel(
        user_id=str(current_user.id),
        user_name=current_user.full_name or current_user.email,
        user_role=current_user.role,
        module="business_case",
        page=f"projects/{project_id}/business-case",
        action="update_manual_values",
        after_state=str(body.model_dump()),
        project_id=project_id,
        timestamp=datetime.now(timezone.utc)
    ))

    await db.commit()
    return {"ok": True, "project_id": project_id}


@router.patch("/projects/{project_id}/business-case/upload/venta")
async def validar_presupuesto_venta(
    project_id: str,
    current_user: UserModel = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session)
):
    """Marca el presupuesto de venta como validado."""
    if project_id in PROYECTOS_PROTEGIDOS:
        raise HTTPException(status_code=403, detail=f"El proyecto {project_id} está protegido.")
    
    result = await db.execute(
        select(BusinessCaseModel).where(BusinessCaseModel.project_id == project_id)
    )
    bc = result.scalar_one_or_none()
    if not bc:
        raise HTTPException(status_code=404, detail="Caso de Negocio no encontrado")
    
    bc.venta_excel_validado = True
    bc.updated_at = datetime.now(timezone.utc)
    await db.commit()
    return {"ok": True}


@router.patch("/projects/{project_id}/business-case/upload/costo")
async def validar_presupuesto_costo(
    project_id: str,
    current_user: UserModel = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session)
):
    """Marca el presupuesto de costo como validado."""
    if project_id in PROYECTOS_PROTEGIDOS:
        raise HTTPException(status_code=403, detail=f"El proyecto {project_id} está protegido.")
    
    result = await db.execute(
        select(BusinessCaseModel).where(BusinessCaseModel.project_id == project_id)
    )
    bc = result.scalar_one_or_none()
    if not bc:
        raise HTTPException(status_code=404, detail="Caso de Negocio no encontrado")
    
    bc.costo_excel_validado = True
    bc.updated_at = datetime.now(timezone.utc)
    await db.commit()
    return {"ok": True}


@router.post("/projects/{project_id}/business-case/upload/venta")
async def upload_presupuesto_venta(
    project_id: str,
    file: UploadFile = File(...),
    trm: Optional[float] = Form(None),
    current_user: UserModel = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session)
):
    """Sube presupuesto de venta, detecta total automáticamente y valida contra el valor manual (Paso 2)."""
    if project_id in PROYECTOS_SIN_PARSER_GENERICO:
        raise HTTPException(status_code=403, detail=f"El proyecto {project_id} utiliza un parser especializado.")

    # 1. Obtener el Caso de Negocio para tener los valores manuales
    result = await db.execute(
        select(BusinessCaseModel).where(BusinessCaseModel.project_id == project_id)
    )
    bc = result.scalar_one_or_none()
    if not bc:
        raise HTTPException(status_code=404, detail="Caso de Negocio no encontrado. Primero guarde los valores manuales en el PASO 1.")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        total_venta = detectar_total_venta(wb)
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error procesando Excel: {str(e)}")

    es_usd = await es_proyecto_usd(project_id, db)
    moneda_val = "USD" if es_usd else "COP"
    
    effective_trm = trm
    if not effective_trm and bc and bc.usd_rate:
        effective_trm = bc.usd_rate
        
    total_excel_converted = total_venta
    if es_usd and effective_trm:
        total_excel_converted = total_venta * effective_trm

    # 2. VALIDACIÓN: ±5% contra el valor manual
    valor_manual = float(bc.venta_monto_manual or 0.0)
    tolerancia = valor_manual * 0.05
    diferencia = abs(total_excel_converted - valor_manual)

    if diferencia > tolerancia:
        raise HTTPException(
            status_code=400,
            detail=(
                f"El total del Excel ({total_excel_converted:,.2f}) no coincide "
                f"con el valor manual ({valor_manual:,.2f}). "
                f"Diferencia: {diferencia:,.2f} (tolerancia máxima: ±{tolerancia:,.2f})"
            )
        )

    # 3. Guardar valores y marcar como validado
    bc.valor_oferta_total = total_excel_converted
    bc.presupuesto_venta_cargado = True
    bc.venta_excel_validado = True
    bc.moneda = moneda_val
    if es_usd and effective_trm:
        bc.usd_rate = effective_trm
    bc.updated_at = datetime.now(timezone.utc)

    # 4. Guardar archivo en project_files
    new_file = ProjectFileModel(
        project_id=project_id,
        nombre_original=file.filename,
        tipo_mime=file.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        tamano_bytes=len(content),
        categoria="presupuesto_venta",
        archivo=content,
        subido_por=current_user.full_name or current_user.email
    )
    db.add(new_file)

    # 5. Registrar actividad
    db.add(ActivityLogModel(
        user_id=str(current_user.id),
        user_name=current_user.full_name or current_user.email,
        user_role=current_user.role,
        module="business_case",
        page=f"projects/{project_id}/business-case",
        action=f"upload_venta: {file.filename}",
        field_name="valor_oferta_total",
        after_state=str(total_excel_converted),
        project_id=project_id,
        timestamp=datetime.now(timezone.utc),
    ))

    await db.commit()
    return {"total_venta": total_excel_converted, "archivo": file.filename, "ok": True}


@router.post("/projects/{project_id}/business-case/upload/costo")
async def upload_presupuesto_costo(
    project_id: str,
    file: UploadFile = File(...),
    trm: Optional[float] = Form(None),
    current_user: UserModel = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session)
):
    """Sube presupuesto de costo, detecta total automáticamente y valida contra el valor manual (Paso 2)."""
    if project_id in PROYECTOS_SIN_PARSER_GENERICO:
        raise HTTPException(status_code=403, detail=f"El proyecto {project_id} utiliza un parser especializado.")

    # 1. Obtener el Caso de Negocio para tener los valores manuales
    result = await db.execute(
        select(BusinessCaseModel).where(BusinessCaseModel.project_id == project_id)
    )
    bc = result.scalar_one_or_none()
    if not bc:
        raise HTTPException(status_code=404, detail="Caso de Negocio no encontrado. Primero guarde los valores manuales en el PASO 1.")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        total_costo = detectar_total_costo(wb)
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error procesando Excel: {str(e)}")

    es_usd = await es_proyecto_usd(project_id, db)
    moneda_val = "USD" if es_usd else "COP"
    
    effective_trm = trm
    if not effective_trm and bc and bc.usd_rate:
        effective_trm = bc.usd_rate

    total_excel_converted = total_costo
    if es_usd and effective_trm:
        total_excel_converted = total_costo * effective_trm

    # 2. VALIDACIÓN: ±5% contra el valor manual
    valor_manual = float(bc.costo_monto_manual or 0.0)
    tolerancia = valor_manual * 0.05
    diferencia = abs(total_excel_converted - valor_manual)

    if diferencia > tolerancia:
        raise HTTPException(
            status_code=400,
            detail=(
                f"El total del Excel ({total_excel_converted:,.2f}) no coincide "
                f"con el valor manual ({valor_manual:,.2f}). "
                f"Diferencia: {diferencia:,.2f} (tolerancia máxima: ±{tolerancia:,.2f})"
            )
        )

    # 3. Guardar valores y marcar como validado
    bc.costo_total = total_excel_converted
    bc.costo_total_sin_fin = total_excel_converted
    bc.costo_total_con_fin = total_excel_converted
    bc.presupuesto_costo_cargado = True
    bc.costo_excel_validado = True
    bc.moneda = moneda_val
    if es_usd and effective_trm:
        bc.usd_rate = effective_trm
    bc.updated_at = datetime.now(timezone.utc)

    # 4. Guardar archivo en project_files
    new_file = ProjectFileModel(
        project_id=project_id,
        nombre_original=file.filename,
        tipo_mime=file.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        tamano_bytes=len(content),
        categoria="presupuesto_costo",
        archivo=content,
        subido_por=current_user.full_name or current_user.email
    )
    db.add(new_file)

    # 5. Registrar actividad
    db.add(ActivityLogModel(
        user_id=str(current_user.id),
        user_name=current_user.full_name or current_user.email,
        user_role=current_user.role,
        module="business_case",
        page=f"projects/{project_id}/business-case",
        action=f"upload_costo: {file.filename}",
        field_name="costo_total",
        after_state=str(total_excel_converted),
        project_id=project_id,
        timestamp=datetime.now(timezone.utc),
    ))

    await db.commit()
    return {"total_costo": total_excel_converted, "archivo": file.filename, "ok": True}


# ══════════════════════════════════════════════════════════════════════════════
# AI PARSER & CATEGORIZATION FOR EXCEL DETAILS
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/projects/{project_id}/business-case/parse-detail")
async def parse_business_case_detail(
    project_id: str,
    body: ParseDetailRequest,
    current_user: UserModel = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session)
):
    """Parsea el Excel de presupuesto y clasifica las líneas detalladas mediante Claude."""
    if not HAS_ANTHROPIC:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Servicio de IA no disponible. Instalar: pip install anthropic"
        )
    import os
    import json
    from io import BytesIO
    import openpyxl
    
    if project_id == 'lyra-carsan-oe2000':
        raise HTTPException(status_code=403, detail="Proyecto protegido. Operación rechazada.")
    tipo = body.tipo.lower().strip()
    if tipo not in ('venta', 'costo'):
        raise HTTPException(status_code=400, detail="El tipo debe ser 'venta' o 'costo'.")

    # 1. Fetch macro total from business_case
    result_bc = await db.execute(
        select(BusinessCaseModel).where(BusinessCaseModel.project_id == project_id)
    )
    bc = result_bc.scalar_one_or_none()
    if not bc:
        raise HTTPException(status_code=404, detail="El Caso de Negocio no existe para este proyecto.")

    macro_total = 0.0
    if tipo == "venta":
        macro_total = float(bc.venta_monto_manual or bc.valor_oferta_total or 0.0)
    else:
        macro_total = float(bc.costo_monto_manual or bc.costo_total or 0.0)

    if macro_total <= 0:
        raise HTTPException(
            status_code=400,
            detail=f"El total macro de {tipo} para el proyecto es 0 o no ha sido configurado."
        )

    # 2. Get latest uploaded excel file
    cat = "presupuesto_venta" if tipo == "venta" else "presupuesto_costo"
    file_result = await db.execute(
        select(ProjectFileModel)
        .where(
            and_(
                ProjectFileModel.project_id == project_id,
                ProjectFileModel.categoria == cat
            )
        )
        .order_by(ProjectFileModel.id.desc())
    )
    excel_file = file_result.scalar_one_or_none()
    if not excel_file:
        raise HTTPException(
            status_code=404,
            detail=f"No se encontró un archivo de presupuesto de {tipo} cargado para este proyecto."
        )

    # 3. Read active sheet with openpyxl (up to 200 rows)
    try:
        wb = openpyxl.load_workbook(BytesIO(excel_file.archivo), data_only=True)
        sheet = wb.active
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Error al abrir el archivo Excel de presupuesto: {str(e)}"
        )

    extracted_rows = []
    for r_idx in range(1, min(sheet.max_row + 1, 201)):
        row_cells = {}
        has_text = False
        has_number = False
        for c_idx in range(1, min(sheet.max_column + 1, 15)):
            cell = sheet.cell(row=r_idx, column=c_idx)
            val = cell.value
            if val is not None:
                cell_ref = cell.coordinate
                if isinstance(val, (int, float)):
                    has_number = True
                    row_cells[cell_ref] = float(val)
                elif isinstance(val, str):
                    val_s = val.strip()
                    if val_s:
                        has_text = True
                        row_cells[cell_ref] = val_s
                else:
                    row_cells[cell_ref] = str(val)
        if has_text and has_number:
            extracted_rows.append({
                "row": r_idx,
                "cells": row_cells
            })

    if not extracted_rows:
        raise HTTPException(
            status_code=400,
            detail="No se encontraron filas con conceptos y montos válidos en las primeras 200 filas del Excel."
        )

    # 4. Call Anthropic Claude-Sonnet
    prompt_content = f"""Eres un experto financiero en control de presupuestos para PC Mejía.
Analiza la siguiente lista de filas extraídas de un archivo Excel de presupuesto de tipo '{tipo}'.
El total macro esperado para este presupuesto es: {macro_total}.

Filas del Excel en formato JSON:
{json.dumps(extracted_rows, ensure_ascii=False, indent=2)}

Instrucciones:
1. Revisa cada fila. Filtra filas de encabezados, subtotales duplicados, IDs o filas vacías. Quédate solo con las líneas de detalle (ítems individuales del presupuesto).
2. Para cada ítem identificado:
   - Extrae el concepto principal (descripción).
   - Extrae su valor monetario total (asegúrate de que sea el total de la línea y no un precio unitario o cantidad).
   - Asigna el ítem a una de estas EXACTAS 4 categorías matching con la empresa:
     * 'Suministro': Equipos, materiales, cables, suministros físicos, etc.
     * 'Mano de Obra': Personal, cuadrillas, técnicos, ingenieros instaladores, etc.
     * 'Administración': Gastos administrativos, ingeniería de diseño, seguros, viáticos, permisos, transporte, etc.
     * 'Otros': Cualquier ítem que no encaje en las anteriores.
   - Registra el cell reference de origen (ej. "B12" o "D4") donde se encuentra la descripción o el valor.

3. Suma el valor de todos los ítems clasificados. Compara la suma con el macro total: {macro_total}.
   El total de los ítems detallados debe estar en un rango de tolerancia del ±10% respecto a {macro_total}.

Retorna tu respuesta estrictamente como un objeto JSON con el siguiente formato, sin explicaciones ni markdown fuera del JSON:
{{
  "items": [
    {{
      "concepto": "Descripción del ítem",
      "valor": 123456.78,
      "categoria": "Suministro",
      "fuente_excel": "B12"
    }}
  ],
  "total_sum": 1234567.89,
  "macro_total": {macro_total},
  "difference_percent": 0.05,
  "is_valid": true,
  "validation_message": "Suma detallada coincide dentro del margen de tolerancia."
}}
"""

    try:
        # Initialize Anthropic. It will pick up ANTHROPIC_API_KEY from os.environ
        client = anthropic.Anthropic()
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4000,
            temperature=0,
            system="Responde únicamente con un objeto JSON válido según las instrucciones.",
            messages=[
                {"role": "user", "content": prompt_content}
            ]
        )
        
        response_text = response.content[0].text.strip()
        if response_text.startswith("```json"):
            response_text = response_text[7:]
        if response_text.endswith("```"):
            response_text = response_text[:-3]
        response_text = response_text.strip()
        
        ai_data = json.loads(response_text)
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error en la clasificación con IA (Claude): {str(e)}"
        )

    items_to_save = ai_data.get("items", [])
    total_sum = ai_data.get("total_sum", 0.0)

    # Secondary check in Python
    calc_diff = abs(total_sum - macro_total) / macro_total if macro_total > 0 else 1.0
    if calc_diff > 0.10:
        return {
            "status": "warning",
            "is_valid": False,
            "macro_total": macro_total,
            "total_sum": total_sum,
            "difference_percent": round(calc_diff * 100, 2),
            "message": f"La suma de los detalles ({total_sum:,.2f}) difiere en {calc_diff * 100:.2f}% del macro total ({macro_total:,.2f}), excediendo el 10% permitido. Ajuste el Excel y reintente.",
            "items_preview": items_to_save[:5]
        }

    # Save items: delete existing for project_id and tipo
    await db.execute(
        delete(BusinessCaseDetailModel)
        .where(
            and_(
                BusinessCaseDetailModel.project_id == project_id,
                BusinessCaseDetailModel.tipo == tipo
            )
        )
    )
    await db.flush()

    saved_items = []
    for item in items_to_save:
        category = item.get("categoria", "Otros").strip()
        if category not in ("Suministro", "Mano de Obra", "Administración", "Otros"):
            category = "Otros"
            
        db_detail = BusinessCaseDetailModel(
            id=str(uuid.uuid4()),
            project_id=project_id,
            tipo=tipo,
            categoria=category,
            concepto=item.get("concepto", "Desconocido")[:500],
            valor=float(item.get("valor", 0.0)),
            moneda=bc.moneda or 'COP',
            fuente_excel=str(item.get("fuente_excel", ""))[:500],
            creado_por_ia=1,
            created_at=datetime.now(timezone.utc)
        )
        db.add(db_detail)
        saved_items.append(db_detail)

    await db.commit()

    return {
        "status": "success",
        "is_valid": True,
        "macro_total": macro_total,
        "total_sum": total_sum,
        "difference_percent": round(calc_diff * 100, 2),
        "message": f"Se clasificaron y guardaron {len(saved_items)} líneas de detalle exitosamente.",
        "items_count": len(saved_items)
    }


@router.get("/projects/{project_id}/business-case/details", response_model=List[BusinessCaseDetailResponse])
async def get_business_case_details(
    project_id: str,
    db: AsyncSession = Depends(get_db_session)
):
    """Retorna todas las líneas de detalle del Caso de Negocio clasificadas por la IA."""
    result = await db.execute(
        select(BusinessCaseDetailModel)
        .where(BusinessCaseDetailModel.project_id == project_id)
        .order_by(BusinessCaseDetailModel.created_at.asc())
    )
    details = list(result.scalars().all())
    return details

