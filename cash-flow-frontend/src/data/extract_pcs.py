import openpyxl, json, re, sys
from datetime import datetime, date

wb = openpyxl.load_workbook(
    'Seguimiento a Proyectos PC Solar_SFR.xlsx',
    data_only=True,
)

LABEL_MAP = {
    'nombre del proyecto': 'nombre_proyecto',
    'nombre del contrato': 'nombre_contrato',
    'codigo del proyecto': 'codigo_proyecto',
    'cliente': 'cliente',
    'gerente del proyecto - cliente': 'gerente_proyecto_cliente',
    'administrador del contrato - cliente': 'administrador_contrato_cliente',
    'interventor': 'interventor',
    'director de proyectos': 'director_proyectos',
    'ingeniero residente': 'ingeniero_residente',
    'supervisor': 'supervisor',
    'encargado': 'encargado',
    'tipo de contrato (obra, suministro, epc, etc.)': 'tipo_contrato',
    'tipo de contrato': 'tipo_contrato',
    'requiere auxilios si / no': 'requiere_auxilios',
    'polizas requeridas': 'polizas_requeridas',
    'multas o penalidades': 'multas_penalidades',
    'alcance:': 'alcance',
    'alcance': 'alcance',
    'localizacion': 'localizacion',
    'fecha de inicio': 'fecha_inicio',
    'fecha de finalizacion contractual': 'fecha_finalizacion_contractual',
    'valor original del contrato': 'valor_original_contrato',
    'porcentaje anticipo': 'porcentaje_anticipo',
    'retencion en garantia': 'retencion_garantia',
    'utilidad proyectada': 'utilidad_proyectada',
    # Seguimiento
    'fecha de terminacion estimada': 'fecha_terminacion_estimada',
    'porcentaje de avance programado a la fecha': 'avance_programado',
    'porcentaje de avance real a la fecha': 'avance_real',
    'modificacion del alcance:': 'modificacion_alcance',
    'ordenes de compra ? (si / no)': 'ordenes_compra',
    'alcance ordenes': 'alcance_ordenes',
    'tiempo ordenes': 'tiempo_ordenes',
    'valor ordenes': 'valor_ordenes',
    'estado facturacion ordenes': 'estado_facturacion_ordenes',
    'desviaciones detectadas': 'desviaciones_detectadas',
    'justificacion de las desviaciones': 'justificacion_desviaciones',
    'valor de los otros': 'valor_otros_adiciones',
    'valor actual del contrato': 'valor_actual_contrato',
    'valor anticipo recibido': 'valor_anticipo_recibido',
    'valor facturado': 'valor_facturado',
    'retenido': 'retenido',
    'amortizacion del anticipo': 'amortizacion_anticipo',
    'valor total ingreso (liquidez)': 'valor_total_ingreso',
    'valor descuentos': 'valor_descuentos',
    'valor pagado': 'valor_pagado',
    'valor por amortizar': 'valor_por_amortizar',
    'costos ejecutados hasta el momento: materiales': 'costos_materiales',
    'costos ejecutados hasta el momento: mano de obra': 'costos_mano_obra',
    'costos ejecutados hasta el momento: administrativos': 'costos_administrativos',
    'costos ejecutados hasta el momento': 'costos_ejecutados_total',
    'utilidad': 'utilidad_actual',
    'utilidad proyectada flujo de caja': 'utilidad_proyectada_fc',
    'necesidades de apoyo': 'necesidades_apoyo',
    'decisiones que deben ser tomadas por la gerencia': 'decisiones_gerencia',
    'observaciones del cliente': 'observaciones_cliente',
    'identificacion de riesgos': 'identificacion_riesgos',
    'lecciones aprendidas': 'lecciones_aprendidas',
    'recomendaciones para otros proyectos': 'recomendaciones',
}

def clean_label(s):
    if not isinstance(s, str):
        return ''
    s = s.strip().lower()
    # Remove special chars
    s = s.replace('\u00ad', '').replace('\xa0', ' ')
    s = s.replace('\u00f3', 'o').replace('\u00e9', 'e').replace('\u00ed', 'i')
    s = s.replace('\u00e1', 'a').replace('\u00fa', 'u').replace('\u00f1', 'n')
    s = s.replace('\u00c9', 'e').replace('\u00bf', '')
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def serialize_val(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, float):
        if v == int(v):
            return int(v)
        return round(v, 6)
    if isinstance(v, int):
        return v
    if isinstance(v, str):
        v = v.replace('\xa0', ' ').strip()
        if not v:
            return None
        # Intento de convertir a número si parece moneda (e.g., "$ 1.000.000")
        if any(c in v for c in ['$', '%']):
            clean = v.replace('$', '').replace('%', '').replace(' ', '').replace('.', '').replace(',', '.')
            try:
                # Si era porcentaje, lo convertimos a decimal base 100
                val = float(clean)
                if '%' in v:
                    return round(val / 100, 6)
                return val
            except:
                pass
    return v

def match_label(label):
    # Exact match
    if label in LABEL_MAP:
        return LABEL_MAP[label]
    # Starts-with match (for truncated labels)
    for lbl_key, fld in LABEL_MAP.items():
        if len(label) >= 20 and label.startswith(lbl_key[:25]):
            return fld
        if len(lbl_key) >= 20 and lbl_key.startswith(label[:25]):
            return fld
    return None

projects = []
for idx, sheet_name in enumerate(wb.sheetnames):
    if sheet_name in ['Costos MO', 'Hoja1']:
        continue
    ws = wb[sheet_name]
    proj = {'sheet_name': sheet_name, 'fecha_informe': None, 'group': 'PCS'}

    # Find fecha del informe
    for row_num in range(1, 6):
        b_val = ws[f'B{row_num}'].value
        if isinstance(b_val, str) and 'fecha del informe' in b_val.lower():
            d_val = ws[f'D{row_num}'].value
            proj['fecha_informe'] = serialize_val(d_val)
            break

    # Scan column C for labels
    for row_num in range(1, ws.max_row + 1):
        c_val = ws[f'C{row_num}'].value
        if c_val is None:
            continue
        label = clean_label(c_val)
        if not label:
            continue

        field = match_label(label)
        if field and field not in proj:
            d_val = ws[f'D{row_num}'].value
            proj[field] = serialize_val(d_val)

    # Generate stable and unique ID
    code = proj.get('codigo_proyecto', '')
    if code:
        base_id = 'pcs-' + re.sub(r'[^a-z0-9]', '-', str(code).lower().strip()).strip('-')
    else:
        base_id = 'pcs-' + re.sub(r'[^a-z0-9]', '-', sheet_name.lower().strip()).strip('-')
    
    proj['id'] = f"{base_id}-{idx}"

    projects.append(proj)

# Write output
output_path = 'projectsSolarData.json'
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(projects, f, ensure_ascii=False, indent=2)

print(f'Extracted {len(projects)} PCS projects')
for p in projects:
    keys = [k for k in p if p[k] is not None]
    print(f'  {p["id"]}: {p["sheet_name"]} ({len(keys)} fields)')
